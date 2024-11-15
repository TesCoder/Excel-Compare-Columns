#!/usr/bin/python3
# This program compares two excel documents and outputs a third document that indicates
# matched first & second column values and their file locations, as well as unmatched values and their file locations.

from openpyxl import *
from tkinter.filedialog import askopenfilename
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import time

root = Tk()
root.geometry('604x360+400+200')
root.title("Excel Comparison")


class Compare():
    def __init__(self):
        self.file1 = ""
        self.file2 = ""

    def setFile1(self, file1):
        print("File 1 set")
        self.file1 = file1
        self.change_pic1()
        return self.file1

    def setFile2(self, file2):
        print("File 2 set")
        self.file2 = file2
        return self.file2

    def __str__(self):
        return str(self.file1 + '\t' + self.file2)

    def process(self):
        # Opening up the files
        file1 = load_workbook(self.file1)
        sheet1 = file1.active
        file2 = load_workbook(self.file2)
        sheet2 = file2.active

        # Initialize sets for storing the (first name, last name) for the two files
        set1 = set()
        set2 = set()

        # Iterate through rows of each file and add to set
        for i in range(1, sheet1.max_row + 1):
            fname = sheet1.cell(row=i, column=1).value.lower()
            lname = sheet1.cell(row=i, column=2).value.lower()

            set1.add((fname, lname))

        for i in range(1, sheet2.max_row + 1):
            fname = sheet2.cell(row=i, column=1).value.lower()
            lname = sheet2.cell(row=i, column=2).value.lower()

            set2.add((fname, lname))

        matched = set1 & set2  # set1.intersection(set2)
        unmatched1 = set1.difference(set2)  # set1 - set2
        unmatched2 = set2.difference(set1)  # set2 - set1

        # Output these sets into new excel sheet
        output = Workbook()
        outSheet = output.active

        for item in matched:
            outSheet.append((item[0], item[1], "matched"))

        for item in unmatched1:
            outSheet.append(
                (item[0], item[1], "unmatched", "Only appears in" + self.file1))

        for item in unmatched2:
            outSheet.append(
                (item[0], item[1], "unmatched", "Only appears in" + self.file2))

        output.save("Output.xlsx")
        print("Analyzed and exported")

        C.finalMessage()  # notifies user process is completed

    def change_pic1(self):
        photo1 = PhotoImage(file=r'images/thumbnail_file_clicked.png')
        compose_button.configure(image=photo1)
        compose_button.photo = photo1
        print("updatedbutton1")

    def change_pic2(self):
        photo1 = PhotoImage(file=r'images/thumbnail_file_clicked.png')
        compose_button2.configure(image=photo1)
        compose_button2.photo = photo1
        print("updatedbutton2")

    def finalMessage(self):
        C.change_pic2()
        root.update()  # refreshes UI to update checked box thumbnail
        time.sleep(2)
        messagebox.showinfo("", "Analyzed and exported")

        C.reset()

    def reset(self):
        C.resetOne()
        C.resetTwo()

    def resetOne(self):
        photo1 = PhotoImage(file=r'images/thumbnail_file.png')
        compose_button.configure(image=photo1)
        compose_button.photo = photo1
        print("resetOne")
        root.update()

    def resetTwo(self):
        photo2 = PhotoImage(file=r'images/thumbnail_file.png')
        compose_button2.configure(image=photo2)
        compose_button2.photo = photo2
        print("resetTwo")
        root.update()


C = Compare()


def OpenFile() -> object:
    file1 = filedialog.askopenfilename()

    print("file1 is: ", file1)

    if ".xlsx" not in file1 and file1 != "":
        messagebox.showinfo("", "Incorrect file type.")
        C.resetOne()
        return
    elif ".xlsx" not in file1 and file1 == "":
        C.resetOne()
        return
    else:
        f1 = C.setFile1(file1)
        print("f1", f1)


frame3 = Frame(root, width=200, height=150, background="white")
frame3.grid(row=0, column=1, rowspan=1, columnspan=50, sticky='w')


def OpenFile2() -> object:
    file2 = filedialog.askopenfilename()

    print("file2 is: ", file2)
    
    if ".xlsx" not in file2 and file2 != "":
        messagebox.showinfo("", "Incorrect file2 type.")
        C.resetTwo()
        return
    elif ".xlsx" not in file2 and file2 == "":
        C.resetTwo()
        return
    else:
        f2 = C.setFile2(file2)
        print("f2", f2)

    if C.file1 is not None and f2 is not None:
        C.process()


prof_img = PhotoImage(file=r'images/background.png')
file1image1 = PhotoImage(file=r'images/thumbnail_file.png')
file1image2 = PhotoImage(file=r'images/thumbnail_file.png')

lbl1 = Label(frame3, image=prof_img, compound=TOP)
lbl1.grid(rowspan=10, columnspan=40, column=0, row=0)

compose_button = Button(frame3, text="Select File 1",
                        image=file1image1, command=OpenFile)
compose_button.grid(column=17, row=5)

compose_button2 = Button(frame3, text="Select File 2",
                         image=file1image2, command=OpenFile2)
compose_button2.grid(column=27, row=5)

root.mainloop()
